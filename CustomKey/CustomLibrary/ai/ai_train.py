import abc
import json
import os
import pickle

import openpyxl as openpyxl
import urllib3
from requests import Session

# 禁用SSL警告
urllib3.disable_warnings()


class AiTrain(metaclass=abc.ABCMeta):
    """日志打入态势，根据告警找到原始日志id，根据id找到原始日志，在AI组提供excel中区别出和没出告警的日志"""
    log_ids_pkl = ''
    warn_logs_pkl = ''
    origin_excel = ''
    new_warn_excel = ''
    new_no_warn_excel = ''

    def __init__(self, domain, cookie):
        """
        :param url: 平台地址
        :param redis_host: redis地址
        :param user: 登录账号
        :param pwd: 登录密码
        """
        self.domain = domain
        self.session = Session()
        self.session.verify = False
        if isinstance(cookie, str):
            self.session.headers = {'cookie': cookie}
        else:
            self.session.cookie = cookie

    def make_url(self, url):
        return f'{self.domain}{url}'

    def get_warn_ids(self):
        url = self.make_url('/maxs/rv/alarm/pageAlarm')
        query = {
            "pageIndex": 1,
            "pageSize": 100,
            "handleStatus": None,
            "riskLevel": None,
            "responseCode": None,
            "alarmType": None,
            "victimArea": None,
            "victimCity": None,
            "victimDomain": None,
            "sourcePlatform": None,
            "assetGroup": None,
            "attackArea": None,
            "attackCity": None,
            "victimIp": None,
            "attackIp": None,
            "attackResult": None,
            "direction": None,
            "killChain": None,
            "isIntel": None,
            "assetKey": None,
            "aiTagChecked": None,
            "startTime": "2023-05-01 00:00:00",
            "endTime": "2023-05-31 00:00:00"
        }

        result = []
        for page in range(1, 1000):
            query.update(pageIndex=page)
            res = self.session.post(url, json=query).json()
            warn_ids = [i['SNOW_ID'] for i in res['data']['rowData']]
            result.extend(warn_ids)
            if res['data']['lastPage']:
                break

        return set(result)

    def get_warn_logs(self, warn_snow_id):
        """获取产生告警的标准日志"""
        url = self.make_url(f'/maxs/rv/alarm/detail/getAlarmDetailById?snowId={warn_snow_id}&esIndex=maxs_alarm_202305')
        res = self.session.get(url).json()
        return res['data']['PARENTUUID_S'].split(',')

    def get_logs_ids(self):
        """获取所有产生告警的日志id"""
        pkl_file = self.log_ids_pkl
        if os.path.exists(pkl_file):
            with open(pkl_file, 'rb') as f:
                return pickle.load(f)

        warn_ids = self.get_warn_ids()
        result = []
        for wid in warn_ids:
            warn_log_ids = self.get_warn_logs(wid)
            result.extend(warn_log_ids)

        r = set(result)
        with open(pkl_file, 'wb') as f:
            pickle.dump(r, f)
            print('save log ids pkl, ok')

        return r

    def get_warn_raw_logs(self):
        logs_ids = self.get_logs_ids()
        standard_logs = self.get_standard_logs(logs_ids)
        return standard_logs

    @staticmethod
    def slice_list(lst, size):
        return [lst[i:i + size] for i in range(0, len(lst), size)]

    def get_raw_msgs_by_id(self, log_snow_ids):
        q = ' OR '.join(log_snow_ids)
        query = {
            "queryTime": "LOG_TIME:[NOW-1YEAR TO NOW]",
            "queryType": "event",
            "q": f"SNOW_ID:({q})",
            "SRC_IP": "",
            "SRC_PORT": "",
            "DST_IP": "",
            "DST_PORT": "",
            "EVENT_NAME": "",
            "EVENT_TYPE": "",
            "selectCondition": "[]",
            "selectIndex": "",
            "start": 1,
            "isAllIndex": False,
            "rows": len(log_snow_ids)
        }

        url = self.make_url('/maxs/rv/fullTextSearch/search')
        res = self.session.post(url, json=query).json()
        msgs = []
        for i in json.loads(res['data']['resultJSON']):
            msg = i['RAW_MSG'].replace('Aug', 'May')
            msgs.append(msg)

        return msgs

    def divide_log(self):
        """读取原始日志excel，分别写入出、没出告警的excel"""
        workbook = openpyxl.load_workbook(self.origin_excel)
        # 获取工作表
        sheet = workbook.active
        warn_logs = set(self.get_warn_raw_logs())
        # 逐行读取数据

        # 创建一个新的Excel文件，写入出告警的日志
        new_wb1 = openpyxl.Workbook()
        new_sheet1 = new_wb1.active

        # 创建一个新的Excel文件，写入未出告警日志
        new_wb2 = openpyxl.Workbook()
        new_sheet2 = new_wb2.active
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        new_sheet1.append(first_row)
        new_sheet2.append(first_row)

        for row in sheet.iter_rows(values_only=True, min_row=2):
            if row[9] in warn_logs:
                new_sheet1.append(row)
            else:
                new_sheet2.append(row)

        # 保存新的Excel文件
        new_wb1.save(self.new_warn_excel)
        new_wb2.save(self.new_no_warn_excel)

        # 关闭工作簿
        workbook.close()
        new_wb1.close()
        new_wb2.close()

    def get_standard_logs(self, log_ids):
        """获取产出告警的原始日志"""
        data = list()
        pkl_file = self.warn_logs_pkl
        if os.path.exists(pkl_file):
            with open(pkl_file, 'rb') as f:
                return pickle.load(f)

        for hundred_ids in self.slice_list(list(log_ids), 100):
            msg_list = self.get_raw_msgs_by_id(hundred_ids)
            data.extend(msg_list)

        with open(pkl_file, 'wb') as f:
            pickle.dump(data, f)

        return data


class AiTrainLvMeng(AiTrain):
    log_ids_pkl = 'log_ids_lm.pkl'
    warn_logs_pkl = 'warn_logs_lm.pkl'
    origin_excel = '/Users/cza/project/yaxin/智能降噪/绿盟日志.xlsx'
    new_warn_excel = '绿盟日志-出告警的.xlsx'
    new_no_warn_excel = '绿盟日志-没出告警的.xlsx'


class AiTrainQiming(AiTrainLvMeng):
    """启明星辰ai训练数据处理"""
    log_ids_pkl = 'log_ids_qmxc.pkl'
    warn_logs_pkl = 'warn_logs_qmxc.pkl'
    origin_excel = '/Users/cza/project/yaxin/智能降噪/启明星日志.xlsx'
    new_warn_excel = '启明星辰日志-出告警的.xlsx'
    new_no_warn_excel = '启明星辰日志-没出告警的.xlsx'


if __name__ == '__main__':
    domain = 'https://192.168.112.124:8686'
    ck = 'ssa_jwt=eyJhbGciOiJSUzI1NiJ9.eyJpYXQiOjE2OTUxMDc5OTIsImp0aSI6IjE3MDQwMzI1NDUxNDIyODgzODUiLCJ0b2tlblRpbWVvdXQiOjE4MDAsInR5cGUiOiJicm93c2VyIiwic3ViIjoiMTcwMTg2MDU3OTc5MTAxMTg0MiIsInBsYXRmb3JtTWciOjAsInRlbmFudElkIjotMSwidXNlcm5hbWUiOiJhZG1pbjMiLCJuYW1lIjoiYWZ0ZXJzY2hvb2wiLCJ1c2VyX3R5cGUiOjF9.Xp3GNuSealg8ybq3_Yw6ht2BkTjijui5FlDT3KlK7kvdmkF840CXVOykVdBPq5luGifzEWotr1gM37FKZhq23Ky2_yOFf2epmoBIbXl0jfs4ZL5i-3Jn-wM1FC5rnvqbW-ugkmqy8qm4HdGUaD2fJNj7q77LagIvWxyHiP2ICqQ8Jia0u9EtP6nylcp2F3g0sNJjuWcijnd1AMQ6Ac5MKNcyB3kJdmJY8LTMUUnRt3HgbbvLj8khElGywJ9Q8cvBbVhghZKqUQqH7AUDg8kPvOXZidw50wafyRuxeWa8Z835KhabfctO9uRThk7gZ-5qhvVvKf4EW58nFVLacY1wEA'
    handler = AiTrainLvMeng(domain, ck)
    handler.divide_log()
    # handler.get_logs_ids()
